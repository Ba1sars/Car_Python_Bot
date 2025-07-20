import sqlite3
import openpyxl
from car_bot.config import DB_NAME, EXCEL_TABLE_NAME


FULL_DB_NAME = "car_bot/database/" + DB_NAME
FULL_EXCEL_TABLE_NAME = "car_bot/database/" + EXCEL_TABLE_NAME


def parsing_data():
    connect = sqlite3.connect(FULL_DB_NAME)
    cursor = connect.cursor()
    cursor.execute(
        "CREATE TABLE IF NOT EXISTS dashcams "
        "(marque, model, series_with_publish_year, "
        "dashcam, photo_link_V1, photo_link_V2, photo_link_V3)"
    )
    file_to_read = openpyxl.load_workbook(
        FULL_EXCEL_TABLE_NAME,
        data_only=True
    )
    sheet = file_to_read["Новый"]
    for row in range(2, sheet.max_row + 1):
        data = []
        for col in range(1, 8):
            value = sheet.cell(row, col).value
            data.append(value)
            all_empty = True
            for v in data:
                if v is not None:
                    if isinstance(v, str) and v.strip() != '':
                        all_empty = False
                        break
                    elif not isinstance(v, str):
                        all_empty = False
                        break
            if all_empty:
                continue
        cursor.execute(
            "INSERT INTO dashcams VALUES (?, ?, ?, ?, ?, ?, ?);",
            (data[0], data[1], data[2], data[3], data[4], data[5], data[6])
        )
    connect.commit()
    connect.close()


def fetch_factory(marque_option, model_option, series_publish_year_option):
    connect = sqlite3.connect(FULL_DB_NAME)
    cursor = connect.cursor()
    return_array = []
    # fetch marque
    if (
        marque_option == "marque" and
        model_option is None and
        series_publish_year_option is None
    ):
        cursor.execute("SELECT marque FROM dashcams")
        return_array = [row[0] for row in cursor.fetchall()]
    # fetch model
    elif (
        marque_option is not None and
        model_option == "model" and
        series_publish_year_option is None
    ):
        cursor.execute(
            "SELECT model FROM dashcams WHERE marque = ?",
            (marque_option,)
        )
        return_array = [row[0] for row in cursor.fetchall()]
    # fetch series and publish year
    elif (
        marque_option is not None and
        model_option is not None and
        series_publish_year_option == "series_with_publish_year"
    ):
        cursor.execute(
            "SELECT series_with_publish_year FROM dashcams "
            "WHERE marque = ? AND model = ?",
            (marque_option, model_option)
        )
        return_array = [row[0] for row in cursor.fetchall()]
    # fetch dashcam and photos
    elif (
        marque_option is not None and
        model_option is not None and
        series_publish_year_option is not None
    ):
        cursor.execute(
            "SELECT dashcam, photo_link_V1, photo_link_V2, photo_link_V3 "
            "FROM dashcams WHERE marque = ? AND model = ? "
            "AND series_with_publish_year = ?",
            (marque_option, model_option, series_publish_year_option)
        )
        return_array = [row for row in cursor.fetchall()]
    connect.commit()
    connect.close()
    return return_array